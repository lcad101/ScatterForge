"""生成测试样本数据（v3.2，模拟真实测试报表结构）。

生成文件：
1. sample_data/示例_原始报表.xlsx —— 仿 LP G2 报表：
   第 1~16 行元数据、第 17 行表头、第 18 行单位、第 19 行起数据。
   数据分三段：低功率(19~118)/中功率(119~218)/高功率(219~318)，
   便于按系列行范围测试「同电压范围对比各功率系列波动」。
2. sample_data/销售数据2026.xlsx —— 简单表格（保留）。
"""
from __future__ import annotations

import math
import os
import random

from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "..", "sample_data")

HEADERS = ["test no.", "Mains Voltage", "Mains Freq.", "Set Temp.", "Temperature",
           "Voltage RMS", "Current RMS", "Power", "PF", "THD", "Ripple"]
UNITS = ["", "[ V ]", "[ Hz ]", "[ C deg. ]", "[ C deg. ]",
         "[ V ]", "[ A ]", "[ W ]", "", "[ % ]", "[ % ]"]


def make_report_sample(path: str) -> str:
    random.seed(2026)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 1~16 行：元数据前导（仿真实报表）
    ws.cell(row=2, column=15, value="2.1.7.00227 (2025-9-17)")
    ws.cell(row=3, column=15, value="template type: Single v90")
    ws.cell(row=4, column=2, value="Report created date :")
    ws.cell(row=4, column=4, value="2026-07-01 08:59:05")
    ws.cell(row=4, column=9, value="Tester:")
    ws.cell(row=4, column=10, value="Zhengjie Shan")
    ws.cell(row=6, column=2, value="Report created by:")
    ws.cell(row=6, column=4, value="shan.zhengjie@signify.com")
    ws.cell(row=11, column=2, value="Power Source :")
    ws.cell(row=11, column=4, value="PPSC,UPC-1,v4.28")
    ws.cell(row=12, column=2, value="Power Meter:")
    ws.cell(row=12, column=4, value="YOKOGAWA,WT1802")
    ws.cell(row=16, column=2, value="Test profile")

    # 17 行：表头；18 行：单位
    for c, (h, u) in enumerate(zip(HEADERS, UNITS), start=1):
        ws.cell(row=17, column=c, value=h)
        ws.cell(row=18, column=c, value=u)

    # 19 行起：数据（低/中/高功率三段，各 100 行）
    base_power = [10.0, 30.0, 60.0]
    for seg, base in enumerate(base_power):
        for i in range(100):
            r = 19 + seg * 100 + i
            voltage = 198 + (i % 55) * 1.02            # 198~254V 波动
            current = base / 230 * (1 + random.uniform(-0.05, 0.05))
            power = base + random.uniform(-base * 0.15, base * 0.15)
            pf = random.uniform(0.95, 0.99)
            thd = 5 + abs(math.sin(i / 12)) * 15 + random.uniform(-1, 1)
            ripple = random.uniform(2, 8)
            ws.cell(row=r, column=1, value=seg * 100 + i + 1)
            ws.cell(row=r, column=2, value=round(voltage, 2))
            ws.cell(row=r, column=3, value=50 if seg != 0 else 50)
            ws.cell(row=r, column=4, value=-40 if seg == 0 else 25)
            ws.cell(row=r, column=5, value=round(voltage, 1))
            ws.cell(row=r, column=6, value=round(voltage, 2))
            ws.cell(row=r, column=7, value=round(current, 4))
            ws.cell(row=r, column=8, value=round(power, 2))
            ws.cell(row=r, column=9, value=round(pf, 4))
            ws.cell(row=r, column=10, value=round(thd, 2))
            ws.cell(row=r, column=11, value=round(ripple, 2))
    wb.save(path)
    return path


def make_sales_sample(path: str) -> str:
    random.seed(42)
    wb = Workbook()
    ws = wb.active
    ws.title = "销售数据"
    ws.append(["月份", "销售额", "利润率", "成本"])
    for i in range(1, 801):
        sales = 100 + 40 * math.sin(i / 30) + random.uniform(-15, 15)
        profit = 15 + 8 * math.sin(i / 50) + random.uniform(-4, 4)
        cost = sales * (1 - profit / 100) + random.uniform(-3, 3)
        ws.append([i, round(sales, 2), round(profit, 2), round(cost, 2)])
    wb.save(path)
    return path


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    p1 = make_report_sample(os.path.join(OUT_DIR, "示例_原始报表.xlsx"))
    p2 = make_sales_sample(os.path.join(OUT_DIR, "销售数据2026.xlsx"))
    return p1, p2


if __name__ == "__main__":
    print(main())
