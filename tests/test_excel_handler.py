"""ExcelHandler 精确行列范围读取测试（v3.2，无表头假设）。"""
import openpyxl
import pytest

from src.core.exceptions import FileMissingException
from src.services.excel_handler import ExcelHandler, numeric_series_points


def make_sample(path, header_row=17, nrows=120, ncols=4):
    """表头在第 header_row 行，数据从 header_row+1 行起。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None] * ncols)  # row1
    # 在 header_row 写表头
    for c, h in enumerate(["X", "Y", "Label", "Z"], start=1):
        ws.cell(row=header_row, column=c, value=h)
    for i in range(1, nrows + 1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=i * 2.0)
        ws.cell(row=r, column=3, value=f"row-{i}")
        ws.cell(row=r, column=4, value=i * 3)
    wb.save(path)
    return path


def test_read_range_full(tmp_path):
    p = make_sample(tmp_path / "s.xlsx")
    sl = ExcelHandler().read_range(str(p), start_row=17, end_row=137, start_col="A", end_col="D")
    assert sl.start_row == 17 and sl.end_row == 137
    assert sl.row_count == 121           # 表头行 17 + 120 数据行
    assert sl.headers == ["X", "Y", "Label", "Z"]
    assert sl.rows[0] == ["X", "Y", "Label", "Z"]      # 第 1 行即表头
    assert sl.rows[1][1] == 2.0                          # 第 18 行 Y


def test_read_range_single_row(tmp_path):
    p = make_sample(tmp_path / "s.xlsx")
    sl = ExcelHandler().read_range(str(p), start_row=19, end_row=None, start_col="A", end_col="D")
    assert sl.row_count == 1
    assert sl.rows[0][0] == 2  # 第 19 行 = 数据行 2


def test_numeric_series_points_with_row_range(tmp_path):
    p = make_sample(tmp_path / "s.xlsx")
    sl = ExcelHandler().read_range(str(p), start_row=17, end_row=137, start_col="A", end_col="D")
    # 系列 1：行 19~60；系列 2：行 61~137
    pts1 = numeric_series_points(sl, "A", "B", 19, 60)
    pts2 = numeric_series_points(sl, "A", "B", 61, 137)
    assert len(pts1) == 42
    assert len(pts2) == 77
    assert pts1[0] == (2.0, 4.0)
    assert pts2[0] == (44.0, 88.0)


def test_file_missing(tmp_path):
    with pytest.raises(FileMissingException):
        ExcelHandler().read_range(str(tmp_path / "nope.xlsx"), 1, 1, "A", "A")
