"""导出服务测试（v3.2：Sheet1=raw data + 每张图一个 Sheet + 覆盖备份 + 轴刻度/连线）。"""
import glob
import os

import openpyxl
from openpyxl import load_workbook

from src.services.chart_builder import ChartOptions, SeriesConfig
from src.services.excel_export_service import ChartSpec, ExcelExportService
from src.services.excel_handler import ExcelHandler


def make_slice(tmp_path):
    p = tmp_path / "src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for c, h in enumerate(["X", "Y", "Z"], start=1):
        ws.cell(row=1, column=c, value=h)
    for i in range(1, 151):
        ws.cell(row=1 + i, column=1, value=i)
        ws.cell(row=1 + i, column=2, value=i * 1.5)
        ws.cell(row=1 + i, column=3, value=i * 0.5)
    wb.save(p)
    return ExcelHandler().read_range(str(p), 1, 151, "A", "C")


def make_specs():
    c1 = ChartSpec(
        chart_name="功率 vs 电压",
        options=ChartOptions(title="功率 vs 电压", x_label="电压", y_label="功率",
                             x_major_unit=5, x_minor_unit=1, show_grid=True, connect_line=True),
        series_list=[
            SeriesConfig("低功率", "A", "B", 2, 60, "#FF0000", "circle", 8),
            SeriesConfig("中功率", "A", "B", 61, 120, "#0000FF", "diamond", 6),
        ],
    )
    c2 = ChartSpec(
        chart_name="Pout vs THD",
        options=ChartOptions(title="Pout vs THD", x_label="THD", y_label="Pout",
                             show_grid=True, connect_line=False),
        series_list=[SeriesConfig("系列1", "A", "C", 2, 151, "#00B050", "square", 6)],
    )
    return [c1, c2]


def test_export_multi_sheet(tmp_path):
    sl = make_slice(tmp_path)
    out = tmp_path / "out.xlsx"
    ExcelExportService().export(sl, make_specs(), str(out))

    wb = load_workbook(str(out))
    assert wb.sheetnames == ["raw data", "功率 vs 电压", "Pout vs THD"]
    ws_data = wb["raw data"]
    assert ws_data.cell(1, 1).value == "X"
    assert ws_data.cell(151, 2).value == 150 * 1.5
    chart1 = wb["功率 vs 电压"]._charts[0]
    assert len(chart1.series) == 2
    # 系列 1 的行范围引用：raw data 第 2~60 行
    xref = chart1.series[0].xVal.numRef.f
    assert "raw data" in xref and "$A$2:$A$60" in xref
    assert chart1.x_axis.majorUnit == 5
    assert chart1.x_axis.minorUnit == 1
    assert chart1.x_axis.majorGridlines is not None
    assert chart1.x_axis.minorGridlines is not None
    chart2 = wb["Pout vs THD"]._charts[0]
    assert len(chart2.series) == 1
    wb.close()


def test_overwrite_backup_and_cleanup(tmp_path):
    sl = make_slice(tmp_path)
    out = tmp_path / "out.xlsx"
    svc = ExcelExportService()
    svc.export(sl, make_specs(), str(out))
    svc2 = ExcelExportService()
    svc2.export(sl, make_specs(), str(out))
    assert os.path.isfile(str(out))
    assert not glob.glob(str(tmp_path / "*.tmp.xlsx"))
    assert not glob.glob(str(tmp_path / "*.backup_*"))


def test_cancel_export(tmp_path):
    import pytest
    from src.core.exceptions import OperationCancelledException
    sl = make_slice(tmp_path)
    svc = ExcelExportService()
    svc.cancel()
    out = tmp_path / "out.xlsx"
    with pytest.raises(OperationCancelledException):
        svc.export(sl, make_specs(), str(out))
    assert not os.path.isfile(str(out))
