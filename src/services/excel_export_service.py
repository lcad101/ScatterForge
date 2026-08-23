"""导出服务（v3.2）：生成全新 Excel 文件，Sheet1=raw data + 每张图一个 Sheet。

- Sheet1 `raw data`：原样复制用户指定范围（真实行列，保留原列位置）。
- 之后每张图一个 Sheet（Sheet 名 = 图表名），内容为原生 ScatterChart。
- 写前备份 → 写 .tmp.xlsx → 校验 → 替换 → 删除备份（含回滚）。
- 禁止修改原用户文件。
"""
from __future__ import annotations

import os
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from typing import Callable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from src.core.exceptions import OperationCancelledException, OperationFailedException
from src.core.stoppable import IStoppable
from src.services.chart_builder import ChartBuilder, ChartOptions, SeriesConfig, check_limit_rules
from src.services.excel_handler import RAW_DATA_SHEET, SheetSlice
from src.models.validators import col_to_index

ProgressCb = Callable[[int], None]
_EXCEED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 浅红


@dataclass
class ChartSpec:
    """一张图表的导出规格。"""
    chart_name: str
    options: ChartOptions
    series_list: list[SeriesConfig]


def safe_sheet_name(name: str, fallback: str = "图表") -> str:
    """Excel Sheet 名合法性：≤31 字符，去除 \\ / ? * [ ] 等非法字符。"""
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", str(name or ""))
    cleaned = cleaned.strip()[:31]
    return cleaned or fallback


class ExcelExportService(IStoppable):
    """生成 + 保存导出文件。"""

    def __init__(self) -> None:
        super().__init__()

    def export(
        self,
        sl: SheetSlice,
        chart_specs: list[ChartSpec],
        target_path: str,
        progress_cb: Optional[ProgressCb] = None,
    ) -> str:
        """完整导出流程，返回最终文件路径。"""
        wb = self.build_workbook(sl, chart_specs, progress_cb)
        self.save_workbook(wb, target_path)
        if progress_cb:
            progress_cb(100)
        return target_path

    def build_workbook(
        self,
        sl: SheetSlice,
        chart_specs: list[ChartSpec],
        progress_cb: Optional[ProgressCb] = None,
    ) -> Workbook:
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = RAW_DATA_SHEET

        # Sheet1：原样复制用户指定范围（保留原列位置），每 1000 行检查取消 + 进度
        total = max(len(sl.rows), 1)
        for i, row in enumerate(sl.rows):
            if i % 1000 == 0:
                if self.should_cancel():
                    raise OperationCancelledException("用户取消导出")
                if progress_cb:
                    progress_cb(int(i / total * 60))
            excel_row = sl.start_row + i
            for j, val in enumerate(row):
                if val is not None:
                    ws_data.cell(row=excel_row, column=sl.start_col + j, value=val)

        # 条件限值规则：超限单元格标红（浅红）+ 追加「限值规则」Sheet
        all_rules = []
        for spec in chart_specs:
            all_rules.extend(spec.options.rules or [])
        if all_rules:
            exceeded = check_limit_rules(sl, all_rules)
            if exceeded:
                xi = col_to_index(all_rules[0].x_col)
                yi = col_to_index(all_rules[0].y_col)
                for p in exceeded:
                    if p["kind"] == "Y":
                        ws_data.cell(row=p["row"], column=yi).fill = _EXCEED_FILL
                    else:
                        ws_data.cell(row=p["row"], column=xi).fill = _EXCEED_FILL
            ws_rules = wb.create_sheet("限值规则")
            ws_rules.append(["X 列", "X 起始", "X 结束", "Y 列", "Y 最小", "Y 最大"])
            for r in all_rules:
                ws_rules.append([r.x_col, r.x_start, r.x_end, r.y_col, r.y_min, r.y_max])
            if exceeded:
                ws_rules.append([])
                ws_rules.append([f"超限点数：{len(exceeded)}"])
                ws_rules.append(["行号", "X 值", "Y 值", "类型"])
                for p in exceeded:
                    ws_rules.append([p["row"], p["x"], p["y"], "X 超限" if p["kind"] == "X" else "Y 超限"])

        # 每张图一个 Sheet（Sheet 名去重，避免与 raw data/限值规则或彼此重名）
        used = {RAW_DATA_SHEET, "限值规则"}
        for idx, spec in enumerate(chart_specs):
            if self.should_cancel():
                raise OperationCancelledException("用户取消导出")
            base = safe_sheet_name(spec.chart_name, f"图表{idx + 1}")
            name, n = base, 2
            while name in used:
                name = f"{base[:27]}({n})"
                n += 1
            used.add(name)
            ws_chart = wb.create_sheet(name)
            chart = ChartBuilder().build(ws_data, spec.series_list, spec.options)
            ws_chart.add_chart(chart, "A1")
            if progress_cb:
                progress_cb(60 + int((idx + 1) / len(chart_specs) * 30))
        return wb

    def save_workbook(self, workbook: Workbook, target_path: str) -> None:
        """写前备份 → .tmp.xlsx → 校验 → 替换 → 删除备份（含回滚）。"""
        backup_path: Optional[str] = None
        tmp_path = target_path + ".tmp.xlsx"
        try:
            if os.path.exists(target_path):
                backup_path = target_path + ".backup_" + datetime.now().strftime("%Y%m%d%H%M%S")
                shutil.copy2(target_path, backup_path)

            if self.should_cancel():
                raise OperationCancelledException("用户取消导出")

            workbook.save(tmp_path)

            if self.should_cancel():
                raise OperationCancelledException("用户取消导出")

            load_workbook(tmp_path).close()  # 校验临时文件可打开

            if os.path.exists(target_path):
                os.remove(target_path)
            os.rename(tmp_path, target_path)

            if backup_path and os.path.exists(backup_path):
                os.remove(backup_path)

        except OperationCancelledException:
            self._rollback(backup_path, tmp_path, target_path)
            raise
        except Exception as exc:  # noqa: BLE001
            self._rollback(backup_path, tmp_path, target_path)
            raise OperationFailedException(f"导出失败: {exc}") from exc

    @staticmethod
    def _rollback(backup_path: Optional[str], tmp_path: str, target_path: str) -> None:
        try:
            if backup_path and os.path.exists(backup_path):
                shutil.copy2(backup_path, target_path)
                os.remove(backup_path)
        except OSError:
            pass
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass
