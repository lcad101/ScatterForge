"""Excel 流式读取（v3.2：精确行列范围，无表头假设）。

架构强制契约：
- read_only 模式，只解析 min_col~max_col 列区间，禁止全量扫描。
- 行列范围 = 真实 Excel 行列号（1 基、含表头所在行），无「表头在第 1 行」假设。
- end_row 为空（None）→ 仅取 start_row 一行；end_col 为空 → 仅取 start_col 一列。
- 实现 IStoppable，供导入流程取消。
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Any, Optional

from openpyxl import load_workbook
import openpyxl.worksheet._reader as _xr

_original_cast = _xr._cast_number

def _safe_cast_number(value):
    """修复 openpyxl 无法处理 'inf'/'nan' 字符串的缺陷。"""
    if isinstance(value, str) and value.lower() in ("inf", "-inf", "nan"):
        return float(value)
    return _original_cast(value)

_xr._cast_number = _safe_cast_number

from src.core.exceptions import FileMissingException, OperationCancelledException
from src.core.stoppable import IStoppable
from src.models.validators import col_to_index, index_to_col

RAW_DATA_SHEET = "raw data"


@dataclass
class SheetSlice:
    """精确行列范围切片。rows 与 headers 均按 start_col..end_col 顺序对齐。"""
    start_row: int
    end_row: int
    start_col: int  # 1 基列索引（源文件列位置，原样保留）
    end_col: int
    headers: list[str] = field(default_factory=list)  # 范围第 1 行（用于列下拉）
    rows: list[list[Any]] = field(default_factory=list)  # 范围内所有行（含表头行）

    @property
    def column_letters(self) -> list[str]:
        return [index_to_col(i) for i in range(self.start_col, self.end_col + 1)]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    def row_index(self, excel_row: int) -> int:
        """真实 Excel 行号 → rows 列表索引。"""
        return excel_row - self.start_row


def _norm_col(col: str | int) -> int:
    return col_to_index(col) if isinstance(col, str) else int(col)


class ExcelHandler(IStoppable):
    """按精确行列范围流式读取 Excel。"""

    def __init__(self) -> None:
        super().__init__()

    def read_range(
        self,
        path: str,
        start_row: int,
        end_row: Optional[int],
        start_col: str | int,
        end_col: Optional[str | int],
        sheet: Optional[str] = None,
    ) -> SheetSlice:
        """读取精确行列范围（真实 Excel 行列号，含表头所在行）。"""
        if not path or not os.path.isfile(path):
            raise FileMissingException(f"文件不存在：{path}")

        sc = _norm_col(start_col)
        ec = sc if end_col is None else _norm_col(end_col)
        if ec < sc:
            ec = sc
        excel_first = start_row
        excel_last = start_row if end_row is None else max(start_row, end_row)

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet else wb.active
            rows: list[list[Any]] = []
            for ri, row in enumerate(ws.iter_rows(min_row=excel_first, max_row=excel_last,
                                                  min_col=sc, max_col=ec, values_only=True)):
                if ri % 1000 == 0 and self.should_cancel():
                    raise OperationCancelledException("用户取消导入")
                rows.append(list(row))
            headers = ExcelHandler._make_headers(rows[0] if rows else [], sc, ec)
            return SheetSlice(start_row=excel_first, end_row=excel_last,
                              start_col=sc, end_col=ec, headers=headers, rows=rows)
        finally:
            wb.close()

    def read_headers(
        self, path: str, start_row: int, start_col: str | int, end_col: Optional[str | int],
        sheet: Optional[str] = None,
    ) -> list[str]:
        sc = _norm_col(start_col)
        ec = sc if end_col is None else _norm_col(end_col)
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet else wb.active
            for row in ws.iter_rows(min_row=start_row, max_row=start_row,
                                    min_col=sc, max_col=ec, values_only=True):
                return ExcelHandler._make_headers(list(row), sc, ec)
            return []
        finally:
            wb.close()

    def count_data_rows(self, path: str, sheet: Optional[str] = None) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet else wb.active
            return max(0, ws.max_row)
        finally:
            wb.close()

    @staticmethod
    def _make_headers(first_row: list, sc: int, ec: int) -> list[str]:
        headers: list[str] = []
        for i, val in enumerate(first_row):
            col_letter = index_to_col(sc + i)
            headers.append(str(val) if val is not None and str(val).strip() != "" else col_letter)
        while len(headers) < (ec - sc + 1):
            headers.append(index_to_col(sc + len(headers)))
        return headers


def numeric_series_points(sl: SheetSlice, x_col: str, y_col: str,
                          row_start: int, row_end: int) -> list[tuple[float, float]]:
    """从切片的指定行范围内抽取 (x, y) 数值点（跳过非数值单元格）。"""
    xi = col_to_index(x_col) - sl.start_col
    yi = col_to_index(y_col) - sl.start_col
    i0 = sl.row_index(row_start)
    i1 = sl.row_index(row_end)
    pts: list[tuple[float, float]] = []
    for i in range(max(0, i0), min(len(sl.rows), i1 + 1)):
        row = sl.rows[i]
        if xi < 0 or yi < 0 or xi >= len(row) or yi >= len(row):
            continue
        try:
            x = float(row[xi])
            y = float(row[yi])
        except (TypeError, ValueError):
            continue
        pts.append((x, y))
    return pts
