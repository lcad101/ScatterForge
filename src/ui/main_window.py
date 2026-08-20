"""主窗口：编排主题树 / 项目列表 / 各视图 / 预览 / 对话框，及数据库持久化。"""
from __future__ import annotations

import os
import sys
from datetime import datetime

from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtWidgets import (
    QApplication, QHBoxLayout, QInputDialog, QLabel, QLineEdit, QListWidget,
    QListWidgetItem, QMainWindow, QMessageBox, QPushButton, QSplitter,
    QTabWidget, QVBoxLayout, QWidget,
)

from src.core.database import SessionLocal
from src.core.exceptions import FileMissingException, OperationCancelledException
from src.models.project import ACTIVE, MISSING, Project
from src.models.series import Series
from src.models.theme import Theme
from src.services.chart_builder import ChartOptions, SeriesConfig
from src.services.excel_export_service import ExcelExportService
from src.services.excel_handler import CHART_VIEW_SHEET, RAW_DATA_SHEET, ExcelHandler, SheetSlice
from src.services.file_probe_service import FileProbeService
from src.ui.dialogs import ExportDialog, ExportImageDialog, default_export_name
from src.ui.views import ConfigView, ImportView, OptionsView
from src.ui.widgets import DragDropThemeTree, PreviewWidget, ProjectListView

DEFAULT_THEME = "我的图表项目"


class ImportWorker(QThread):
    success = Signal(object)  # SheetSlice
    failed = Signal(str)
    cancelled = Signal()

    def __init__(self, path, start_row, end_row, start_col, end_col, read_raw=False, parent=None):
        super().__init__(parent)
        self._handler = ExcelHandler()
        self._args = (path, start_row, end_row, start_col, end_col)
        self._read_raw = read_raw

    def cancel(self):
        self._handler.cancel()

    def run(self):
        try:
            path, sr, er, sc, ec = self._args
            if self._read_raw:
                sl = self._handler.read_raw_data_sheet(path, sr, er, sc, ec)
            else:
                sl = self._handler.read_slice(path, sr, er, sc, ec)
            self.success.emit(sl)
        except OperationCancelledException:
            self.cancelled.emit()
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))


class ProbeWorker(QThread):
    done = Signal()
    failed = Signal(str)

    def run(self):
        session = SessionLocal()
        try:
            FileProbeService().probe(session)
            self.done.emit()
        except Exception as exc:  # noqa: BLE001
            self.failed.emit(str(exc))
        finally:
            session.close()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel 散点图生成器 — ScatterForge")
        self.resize(1280, 800)

        self.db = SessionLocal()
        self._current_project: Project | None = None
        self._source_slice: SheetSlice | None = None
        self._source_basename: str = ""
        self._series: list[SeriesConfig] = []
        self._options = ChartOptions()
        self._import_worker: ImportWorker | None = None

        self._build_ui()
        self._refresh_all()
        self._probe_async()

    # ================= UI 构建 =================
    def _build_ui(self):
        # ---- 左：主题树 ----
        self.tree = DragDropThemeTree()
        self.tree.itemSelectionChanged.connect(self._on_tree_selected)
        self.tree.orderChanged.connect(self._on_theme_reordered)

        add_theme_btn = QPushButton("＋ 新建主题")
        add_theme_btn.clicked.connect(self._add_theme)
        del_theme_btn = QPushButton("🗑 删除主题")
        del_theme_btn.clicked.connect(self._delete_theme)
        theme_btns = QHBoxLayout()
        theme_btns.addWidget(add_theme_btn)
        theme_btns.addWidget(del_theme_btn)

        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(4, 4, 4, 4)
        left_layout.addWidget(QLabel("📁 主题分组"))
        left_layout.addWidget(self.tree, 1)
        left_layout.addLayout(theme_btns)

        # ---- 中：项目列表 ----
        self.search = QLineEdit()
        self.search.setPlaceholderText("🔍 搜索项目名称或路径…")
        self.search.textChanged.connect(lambda _: self._refresh_list())

        self.list = ProjectListView()
        self.list.projectActivated.connect(self._open_project)
        self.list.selectionChangedSig.connect(self._on_selection_changed)

        self.sel_label = QLabel("已选 0 个")
        self.batch_export_btn = QPushButton("批量导出")
        self.batch_export_btn.clicked.connect(self._batch_export)
        self.batch_delete_btn = QPushButton("批量删除")
        self.batch_delete_btn.clicked.connect(self._batch_delete)
        batch_row = QHBoxLayout()
        batch_row.addWidget(self.sel_label)
        batch_row.addWidget(self.batch_export_btn)
        batch_row.addWidget(self.batch_delete_btn)

        mid = QWidget()
        mid_layout = QVBoxLayout(mid)
        mid_layout.setContentsMargins(4, 4, 4, 4)
        mid_layout.addWidget(QLabel("📋 图表项目"))
        mid_layout.addWidget(self.search)
        mid_layout.addWidget(self.list, 1)
        mid_layout.addLayout(batch_row)

        # ---- 右：标签页 ----
        self.tabs = QTabWidget()
        self.start_page = self._build_start_page()
        self.import_view = ImportView()
        self.import_view.applyRequested.connect(self._import)
        self.config_view = ConfigView()
        self.config_view.changed.connect(self._preview_from_state)
        self.config_view.exportRequested.connect(self._export)
        self.options_view = OptionsView()
        self.options_view.changed.connect(self._preview_from_state)
        self.preview = PreviewWidget()

        self.tabs.addTab(self.start_page, "开始")
        self.tabs.addTab(self.import_view, "数据设置")
        self.tabs.addTab(self.config_view, "系列配置")
        self.tabs.addTab(self.options_view, "图表选项")
        self.tabs.addTab(self.preview, "图表预览")

        # ---- 布局 ----
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(left)
        splitter.addWidget(mid)
        splitter.addWidget(self.tabs)
        splitter.setSizes([220, 320, 700])
        self.setCentralWidget(splitter)

        # ---- 菜单 ----
        menubar = self.menuBar()
        file_menu = menubar.addMenu("文件(&F)")
        file_menu.addAction("导入原始数据…", self._goto_import)
        file_menu.addAction("导出图表…", self._export)
        file_menu.addAction("导出为图片…", self._export_image)
        self.recent_menu = file_menu.addMenu("最近打开")
        file_menu.addSeparator()
        file_menu.addAction("退出", self.close)
        help_menu = menubar.addMenu("帮助(&H)")
        help_menu.addAction("关于", self._about)

        # ---- 工具栏 ----
        tb = self.addToolBar("主工具栏")
        tb.setMovable(False)
        tb.addAction("📥 导入原始数据", self._goto_import)
        tb.addAction("📈 导出图表", self._export)
        tb.addAction("📷 导出图片", self._export_image)
        tb.addAction("🔄 刷新状态", self._probe_async)
        tb.addAction("📦 批量导出", self._batch_export)

        self.statusBar().showMessage("就绪")

    def _build_start_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.addWidget(QLabel("📊 最近打开的项目（最近 5 个）"))
        self.recent_list = QListWidget()
        self.recent_list.itemDoubleClicked.connect(self._on_recent_clicked)
        layout.addWidget(self.recent_list, 1)
        row = QHBoxLayout()
        import_btn = QPushButton("📂 导入新数据")
        import_btn.clicked.connect(self._goto_import)
        browse_btn = QPushButton("📁 浏览全部项目")
        browse_btn.clicked.connect(lambda: self.tabs.setCurrentIndex(0))
        row.addWidget(import_btn)
        row.addWidget(browse_btn)
        layout.addLayout(row)
        return page

    # ================= 数据刷新 =================
    def _themes_data(self) -> list[dict]:
        themes = self.db.query(Theme).order_by(Theme.sort_order, Theme.id).all()
        result = []
        for t in themes:
            projects = []
            for p in sorted(t.projects, key=lambda x: x.id):
                projects.append({
                    "id": p.id, "name": p.name, "path": p.source_file_path, "status": p.status,
                })
            result.append({"id": t.id, "name": t.name, "projects": projects})
        return result

    def _refresh_all(self):
        self._refresh_tree()
        self._refresh_list()
        self._refresh_recent()

    def _refresh_tree(self):
        themes = self._themes_data()
        active_id = self._current_project.theme_id if self._current_project else None
        self.tree.set_data(themes, active_theme_id=active_id)

    def _refresh_list(self):
        text = self.search.text().strip().lower()
        projects = []
        for t in self._themes_data():
            for p in t["projects"]:
                if text and text not in p["name"].lower() and text not in p["path"].lower():
                    continue
                projects.append(p)
        self.list.set_projects(projects)

    def _refresh_recent(self):
        self.recent_list.clear()
        recent = (self.db.query(Project)
                  .filter(Project.last_opened_at.isnot(None))
                  .order_by(Project.last_opened_at.desc())
                  .limit(5).all())
        for p in recent:
            mark = "❌ " if p.status == MISSING else ""
            item = QListWidgetItem(f"{mark}{p.name}\n{p.source_file_path}")
            item.setData(Qt.ItemDataRole.UserRole, p.id)
            self.recent_list.addItem(item)
        # 最近打开子菜单
        self.recent_menu.clear()
        for p in recent:
            act = self.recent_menu.addAction(f"{p.name}")
            act.triggered.connect(lambda _=False, pid=p.id: self._open_project_id(pid))

    # ================= 主题操作 =================
    def _default_theme(self) -> Theme:
        theme = self.db.query(Theme).filter(Theme.name == DEFAULT_THEME).first()
        if not theme:
            theme = Theme(name=DEFAULT_THEME, sort_order=0)
            self.db.add(theme)
            self.db.commit()
        return theme

    def _add_theme(self):
        name, ok = QInputDialog.getText(self, "新建主题", "主题名称：")
        if ok and name.strip():
            if self.db.query(Theme).filter(Theme.name == name.strip()).first():
                QMessageBox.warning(self, "提示", "主题名称已存在")
                return
            max_order = max([t.sort_order for t in self.db.query(Theme).all()] or [0])
            self.db.add(Theme(name=name.strip(), sort_order=max_order + 1))
            self.db.commit()
            self._refresh_tree()

    def _delete_theme(self):
        theme_id = self._selected_theme_id()
        if not theme_id:
            return
        theme = self.db.get(Theme, theme_id)
        if not theme:
            return
        if QMessageBox.question(self, "删除主题", f"确定删除主题「{theme.name}」？其下项目也会级联删除（磁盘文件不会被删除）。") != QMessageBox.StandardButton.Yes:
            return
        self.db.delete(theme)
        self.db.commit()
        self._refresh_all()

    def _selected_theme_id(self) -> int | None:
        items = self.tree.selectedItems()
        if not items:
            return None
        tid = items[0].data(0, DragDropThemeTree.THEME_ROLE)
        return tid if tid is not None else None

    def _on_tree_selected(self):
        tid = self._selected_theme_id()
        if tid:
            # 按主题过滤列表（简单实现：刷新列表为当前主题下项目）
            self._refresh_list_by_theme(tid)

    def _refresh_list_by_theme(self, theme_id: int):
        theme = self.db.get(Theme, theme_id)
        if not theme:
            return
        projects = [{"id": p.id, "name": p.name, "path": p.source_file_path, "status": p.status}
                    for p in sorted(theme.projects, key=lambda x: x.id)]
        self.list.set_projects(projects)

    def _on_theme_reordered(self):
        order = self.tree.theme_order()
        for i, tid in enumerate(order):
            theme = self.db.get(Theme, tid)
            if theme and theme.sort_order != i:
                theme.sort_order = i
        self.db.commit()

    # ================= 导入 / 打开 =================
    def _goto_import(self):
        self.tabs.setCurrentWidget(self.import_view)

    def _import(self, path, start_row, end_row, start_col, end_col):
        self._source_basename = os.path.basename(path)
        self._run_import(path, start_row, end_row, start_col, end_col, read_raw=False)

    def _open_project(self, project_id: int):
        self._open_project_id(project_id)

    def _open_project_id(self, project_id: int):
        project = self.db.get(Project, project_id)
        if not project:
            return
        if not os.path.isfile(project.source_file_path):
            project.status = MISSING
            self.db.commit()
            self._refresh_all()
            self._relocate(project)
            return
        self._current_project = project
        project.last_opened_at = datetime.now()
        project.status = ACTIVE
        self.db.commit()
        # 从 raw data 页重新读取（场景 2）
        self._run_import(project.source_file_path,
                         project.raw_start_row, project.raw_end_row,
                         project.raw_start_col_letter, project.raw_end_col_letter,
                         read_raw=True,
                         series=[SeriesConfig(
                             name=s.series_name, x_col=s.x_col_letter, y_col=s.y_col_letter,
                             color=s.color_hex, shape=s.marker_shape, size=s.marker_size,
                         ) for s in project.series_list],
                         options=ChartOptions(
                             title=project.chart_title, x_label=project.x_axis_label,
                             y_label=project.y_axis_label, x_min=project.x_axis_min,
                             x_max=project.x_axis_max, y_min=project.y_axis_min,
                             y_max=project.y_axis_max, show_grid=project.show_grid,
                         ))

    def _run_import(self, path, start_row, end_row, start_col, end_col, read_raw=False,
                    series=None, options=None):
        self.statusBar().showMessage("正在读取数据…")
        self._import_worker = ImportWorker(path, start_row, end_row, start_col, end_col, read_raw, self)
        self._import_worker.success.connect(
            lambda sl, s=series, o=options: self._on_import_done(sl, s, o))
        self._import_worker.failed.connect(self._on_import_failed)
        self._import_worker.cancelled.connect(lambda: self.statusBar().showMessage("导入已取消"))
        self._import_worker.start()

    def _on_import_done(self, sl: SheetSlice, series, options):
        self._source_slice = sl
        # 默认系列
        if series is None:
            letters = sl.column_letters
            x = letters[0] if letters else "A"
            y = letters[1] if len(letters) > 1 else x
            series = [SeriesConfig(name="系列 1", x_col=x, y_col=y, color="#FF0000", shape="circle", size=8)]
        if options is None:
            options = ChartOptions(title=self._source_basename.rsplit(".", 1)[0], show_grid=True)
        self._series = series
        self._options = options

        self.config_view.set_series(self._series)
        self.options_view.set_options(self._options)
        self.import_view.set_max_row(sl.row_count)
        self._preview_from_state()
        self.statusBar().showMessage(f"已加载 {sl.row_count:,} 行数据")
        self.tabs.setCurrentWidget(self.config_view)

    def _on_import_failed(self, msg: str):
        self.statusBar().showMessage("读取失败")
        QMessageBox.critical(self, "读取失败", msg)

    # ================= 预览 / 选项 =================
    def _preview_from_state(self):
        if not self._source_slice:
            return
        self._series = self.config_view.get_series()
        self._options = self.options_view.get_options()
        self.preview.render(self._source_slice, self._series, self._options)

    # ================= 导出 =================
    def _export(self):
        if not self._source_slice:
            QMessageBox.information(self, "提示", "请先导入数据或打开项目")
            return
        if self._current_project and os.path.isfile(self._current_project.source_file_path):
            default_path = self._current_project.source_file_path
        else:
            default_path = default_export_name(self._source_basename)
        dlg = ExportDialog(self._source_slice, self.config_view.get_series(),
                           self.options_view.get_options(), default_path, self)
        if dlg.exec() == ExportDialog.DialogCode.Accepted:
            path = dlg.result_path
            self._persist_project(path)
            self.statusBar().showMessage(f"导出完成：{path}")

    def _persist_project(self, path: str):
        name = os.path.splitext(os.path.basename(path))[0]
        theme = self._default_theme()
        options = self.options_view.get_options()
        series_cfgs = self.config_view.get_series()

        if self._current_project is None:
            project = Project(
                name=name, theme_id=theme.id, source_file_path=path,
                raw_start_row=1,
                raw_end_row=self._source_slice.row_count,
                raw_start_col_letter=self._source_slice.column_letters[0],
                raw_end_col_letter=self._source_slice.column_letters[-1],
                status=ACTIVE, last_opened_at=datetime.now(),
            )
            self.db.add(project)
            self.db.flush()
            self._current_project = project
        else:
            project = self._current_project
            project.source_file_path = path
            project.last_opened_at = datetime.now()
            project.status = ACTIVE

        # 图表元数据
        project.chart_title = options.title
        project.x_axis_label = options.x_label
        project.y_axis_label = options.y_label
        project.x_axis_min = options.x_min
        project.x_axis_max = options.x_max
        project.y_axis_min = options.y_min
        project.y_axis_max = options.y_max
        project.show_grid = options.show_grid

        # 系列（重建）
        for s in list(project.series_list):
            self.db.delete(s)
        self.db.flush()
        for i, cfg in enumerate(series_cfgs):
            self.db.add(Series(
                project_id=project.id, series_name=cfg.name,
                x_col_letter=cfg.x_col, y_col_letter=cfg.y_col,
                color_hex=cfg.color, marker_shape=cfg.shape, marker_size=cfg.size,
                sort_order=i,
            ))
        self.db.commit()
        self._refresh_all()

    def _export_image(self):
        if not self._source_slice:
            QMessageBox.information(self, "提示", "请先导入数据或打开项目")
            return
        self.tabs.setCurrentWidget(self.preview)
        self._preview_from_state()
        name = os.path.splitext(self._source_basename or "散点图")[0] + "_散点图.png"
        dlg = ExportImageDialog(self.preview, name, self)
        dlg.exec()

    # ================= 批量操作（Q2） =================
    def _on_selection_changed(self, n: int):
        self.sel_label.setText(f"已选 {n} 个")

    def _batch_export(self):
        ids = self.list.selected_project_ids()
        if not ids:
            QMessageBox.information(self, "提示", "请先在列表中选中项目（Ctrl/Shift 多选）")
            return
        from PySide6.QtWidgets import QFileDialog
        folder = QFileDialog.getExistingDirectory(self, "选择批量导出的目标文件夹")
        if not folder:
            return
        exported = 0
        for pid in ids:
            p = self.db.get(Project, pid)
            if not p or p.status != ACTIVE:
                continue
            try:
                self._re_export_project(p, folder)
                exported += 1
            except Exception as exc:  # noqa: BLE001
                QMessageBox.warning(self, "批量导出", f"{p.name} 导出失败：{exc}")
        QMessageBox.information(self, "批量导出", f"完成，成功导出 {exported} 个项目到：\n{folder}")

    def _re_export_project(self, project: Project, folder: str):
        h = ExcelHandler()
        sl = h.read_raw_data_sheet(project.source_file_path, project.raw_start_row,
                                   project.raw_end_row, project.raw_start_col_letter,
                                   project.raw_end_col_letter)
        series = [SeriesConfig(name=s.series_name, x_col=s.x_col_letter, y_col=s.y_col_letter,
                               color=s.color_hex, shape=s.marker_shape, size=s.marker_size)
                  for s in project.series_list]
        options = ChartOptions(title=project.chart_title, x_label=project.x_axis_label,
                               y_label=project.y_axis_label, x_min=project.x_axis_min,
                               x_max=project.x_axis_max, y_min=project.y_axis_min,
                               y_max=project.y_axis_max, show_grid=project.show_grid)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = os.path.join(folder, f"{project.name}_{stamp}.xlsx")
        ExcelExportService().export(sl, series, options, target)

    def _batch_delete(self):
        ids = self.list.selected_project_ids()
        if not ids:
            QMessageBox.information(self, "提示", "请先在列表中选中项目")
            return
        if QMessageBox.question(self, "批量删除", f"确定删除 {len(ids)} 个项目记录吗？\n（磁盘文件不会被删除）") != QMessageBox.StandardButton.Yes:
            return
        for pid in ids:
            p = self.db.get(Project, pid)
            if p:
                self.db.delete(p)
        self.db.commit()
        self._refresh_all()
        self.statusBar().showMessage(f"已删除 {len(ids)} 个项目记录")

    # ================= 重定位 / 删除（MISSING） =================
    def _relocate(self, project: Project):
        from PySide6.QtWidgets import QFileDialog
        if QMessageBox.question(self, "文件缺失", f"项目「{project.name}」的导出文件不存在。\n是否重新定位到新文件？") != QMessageBox.StandardButton.Yes:
            return
        path, _ = QFileDialog.getOpenFileName(self, "选择导出文件", "", "Excel 文件 (*.xlsx)")
        if not path:
            return
        # 校验含 raw data 页
        from openpyxl import load_workbook
        try:
            wb = load_workbook(path, read_only=True)
            if RAW_DATA_SHEET not in wb.sheetnames:
                raise FileMissingException(f"所选文件不包含 {RAW_DATA_SHEET} 页")
            wb.close()
        except FileMissingException as exc:
            QMessageBox.warning(self, "校验失败", str(exc))
            return
        except Exception:
            QMessageBox.warning(self, "校验失败", "无法打开所选文件")
            return
        project.source_file_path = path
        project.status = ACTIVE
        self.db.commit()
        self._refresh_all()
        self.statusBar().showMessage(f"已重新定位：{path}")

    # ================= 其他 =================
    def _on_recent_clicked(self, item: QListWidgetItem):
        pid = item.data(Qt.ItemDataRole.UserRole)
        if pid is not None:
            self._open_project_id(pid)

    def _probe_async(self):
        self.statusBar().showMessage("正在后台探活文件状态…")
        self._probe_worker = ProbeWorker(self)
        self._probe_worker.done.connect(self._on_probe_done)
        self._probe_worker.failed.connect(lambda m: self.statusBar().showMessage(f"探活失败：{m}"))
        self._probe_worker.start()

    def _on_probe_done(self):
        self.db.expire_all()
        self._refresh_all()
        self.statusBar().showMessage("就绪")

    def _about(self):
        QMessageBox.about(
            self, "关于",
            "Excel 散点图生成器 — ScatterForge\n\n"
            "Excel 文件生成器/转换器：导入原始数据 → 生成含 raw data + 原生散点图的全新 .xlsx。\n"
            "技术栈：PySide6 · openpyxl · SQLAlchemy · ECharts"
        )

    def closeEvent(self, event):
        self.db.close()
        super().closeEvent(event)


def run_app() -> int:
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    return app.exec()
